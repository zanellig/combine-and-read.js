import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from abc import ABC, abstractmethod

# Abstract Factory
class DataFrameFactory(ABC):
    @abstractmethod
    def create_dataframe(self, file_path: str) -> pd.DataFrame:
        pass

# Concrete Factory
class CSVDataFrameFactory(DataFrameFactory):
    def create_dataframe(self, file_path: str) -> pd.DataFrame:
        return pd.read_csv(file_path)

# Client Code
class ExcelExporter:
    def __init__(self, dataframe_factory: DataFrameFactory):
        self.dataframe_factory = dataframe_factory

    def export_to_excel(self, input_file: str, output_file: str):
        # Create DataFrame
        df = self.dataframe_factory.create_dataframe(input_file)
        
        # Convert DataFrame to Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)  # Write data with headers
            
            # Get the workbook and the first worksheet
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            # Style the first three columns of the header row
            header_font = Font(color="FFFFFF")
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            
            for col in range(1, 4):  # Column 1, 2, and 3
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill

def find_latest_csv(directory: str, base_name: str) -> str:
    # Pattern to match the base file name and capture the counter
    pattern = re.compile(rf"{re.escape(base_name)}(?:\((\d+)\))?\.csv$")
    max_counter = -1
    latest_file = None

    for file_name in os.listdir(directory):
        match = pattern.match(file_name)
        if match:
            counter = int(match.group(1)) if match.group(1) else 0
            if counter > max_counter:
                max_counter = counter
                latest_file = os.path.join(directory, file_name)

    return latest_file

if __name__ == "__main__":
    base_csv_name = "UXPos BackOffice Personas"
    output_directory = os.path.join(os.path.dirname(__file__), '..', '..', 'public', 'output')

    latest_csv_file = find_latest_csv(output_directory, base_csv_name)

    if latest_csv_file:
        csv_factory = CSVDataFrameFactory()
        exporter = ExcelExporter(csv_factory)
        output_xlsx_file = latest_csv_file.replace('.csv', '.xlsx')
        exporter.export_to_excel(latest_csv_file, output_xlsx_file)
        print(f"Converted {latest_csv_file} to {output_xlsx_file}")
    else:
        print("No CSV files found.")
